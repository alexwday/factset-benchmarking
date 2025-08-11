#!/usr/bin/env python3
"""
FactSet Fundamentals Coverage Matrix Generator
Generates a comprehensive comparison matrix showing which fundamental metrics
are available across all monitored banks (Canadian and US banks).
Outputs to Excel with detailed analysis.
"""

import os
import sys
import json
import time
import logging
from datetime import datetime, timedelta
from urllib.parse import quote
from typing import Dict, List, Optional, Any, Tuple
import warnings

import pandas as pd
import yaml
import fds.sdk.FactSetFundamentals
from fds.sdk.FactSetFundamentals.api import metrics_api, fact_set_fundamentals_api
from fds.sdk.FactSetFundamentals.models import *
from fds.sdk.FactSetFundamentals.model.ids_batch_max30000 import IdsBatchMax30000
from fds.sdk.FactSetFundamentals.model.metrics import Metrics
from fds.sdk.FactSetFundamentals.model.periodicity import Periodicity
from fds.sdk.FactSetFundamentals.model.update_type import UpdateType
from fds.sdk.FactSetFundamentals.model.fiscal_period import FiscalPeriod
from fds.sdk.FactSetFundamentals.model.batch import Batch
from fds.sdk.FactSetFundamentals.model.fundamental_request_body import FundamentalRequestBody
from fds.sdk.FactSetFundamentals.model.fundamentals_request import FundamentalsRequest
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Suppress warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(f'coverage_matrix_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Target period for comparison (Q1 2025)
TARGET_FISCAL_YEAR = 2025
TARGET_FISCAL_QUARTER = 1
TARGET_DATE_START = "2025-01-01"
TARGET_DATE_END = "2025-03-31"

def validate_env_vars():
    """Validate required environment variables."""
    required_vars = [
        'API_USERNAME', 'API_PASSWORD', 
        'PROXY_USER', 'PROXY_PASSWORD', 'PROXY_URL',
        'PROXY_DOMAIN'
    ]
    
    missing = [var for var in required_vars if not os.getenv(var)]
    if missing:
        logger.error(f"Missing required environment variables: {', '.join(missing)}")
        sys.exit(1)
    
    logger.info("✅ All required environment variables present")

def setup_ssl_certificate() -> Optional[str]:
    """Setup SSL certificate for API connection."""
    # Check for certificate in input folder first
    input_cert_path = "input/rbc-ca-bundle.cer"
    
    if os.path.exists(input_cert_path):
        logger.info(f"✅ Using SSL certificate from input folder: {input_cert_path}")
        return input_cert_path
    
    # Check for certificate in current directory
    local_cert_path = "rbc-ca-bundle.cer"
    if os.path.exists(local_cert_path):
        logger.info(f"✅ Using SSL certificate from current directory: {local_cert_path}")
        return local_cert_path
    
    # Check environment variable for custom path
    env_cert_path = os.getenv('SSL_CERT_PATH')
    if env_cert_path and os.path.exists(env_cert_path):
        logger.info(f"✅ Using SSL certificate from environment variable: {env_cert_path}")
        return env_cert_path
    
    logger.error("❌ SSL certificate not found!")
    logger.info("Please place the certificate in one of these locations:")
    logger.info("  1. input/rbc-ca-bundle.cer (recommended)")
    logger.info("  2. ./rbc-ca-bundle.cer (current directory)")
    logger.info("  3. Set SSL_CERT_PATH environment variable to certificate path")
    return None

def setup_api_configuration(ssl_cert_path: Optional[str]):
    """Setup FactSet API configuration with proxy and SSL."""
    api_username = os.getenv('API_USERNAME')
    api_password = os.getenv('API_PASSWORD')
    proxy_user = os.getenv('PROXY_USER')
    proxy_password = os.getenv('PROXY_PASSWORD')
    proxy_url = os.getenv('PROXY_URL')
    proxy_domain = os.getenv('PROXY_DOMAIN', 'MAPLE')
    
    # Setup proxy URL with NTLM authentication
    escaped_domain = quote(f"{proxy_domain}\\{proxy_user}")
    quoted_password = quote(proxy_password)
    full_proxy_url = f"http://{escaped_domain}:{quoted_password}@{proxy_url}"
    
    # Configure API client
    configuration = fds.sdk.FactSetFundamentals.Configuration(
        username=api_username,
        password=api_password,
        proxy=full_proxy_url,
        ssl_ca_cert=ssl_cert_path
    )
    
    # Generate authentication token
    configuration.get_basic_auth_token()
    
    logger.info("✅ FactSet API client configured")
    return configuration

def load_config_yaml() -> Dict[str, Any]:
    """Load configuration from config.yaml file."""
    config_path = "example/config.yaml"
    
    if not os.path.exists(config_path):
        logger.error(f"Config file not found: {config_path}")
        sys.exit(1)
    
    with open(config_path, 'r') as f:
        config = yaml.safe_load(f)
    
    logger.info(f"✅ Loaded configuration from {config_path}")
    return config

def get_canadian_and_us_banks(config: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
    """Extract only Canadian and US banks from config."""
    monitored = config.get('monitored_institutions', {})
    
    # Filter for Canadian and US banks only
    canadian_us_banks = {}
    
    for ticker, info in monitored.items():
        bank_type = info.get('type', '')
        if bank_type in ['Canadian_Banks', 'US_Banks', 'US_Regionals']:
            canadian_us_banks[ticker] = info
    
    logger.info(f"📊 Found {len(canadian_us_banks)} Canadian and US banks to analyze")
    return canadian_us_banks

def get_all_available_metrics(api_client) -> Dict[str, List[Dict[str, Any]]]:
    """Fetch all available metrics from the API grouped by category."""
    logger.info("🔍 Fetching all available metrics from FactSet Fundamentals API...")
    
    data_api = metrics_api.MetricsApi(api_client)
    
    # Categories to fetch metrics for
    categories = [
        "INCOME_STATEMENT",
        "BALANCE_SHEET", 
        "CASH_FLOW",
        "RATIOS",
        "FINANCIAL_SERVICES",
        "INDUSTRY_METRICS",
        "PENSION_AND_POSTRETIREMENT",
        "MARKET_DATA",
        "MISCELLANEOUS",
        "DATES"
    ]
    
    all_metrics = {}
    total_metrics = 0
    
    for category in categories:
        try:
            logger.info(f"  📂 Fetching {category} metrics...")
            
            # API call to get metrics for category
            response = data_api.get_fds_fundamentals_metrics(category=category)
            
            if response and hasattr(response, 'data') and response.data:
                metrics_list = []
                for metric in response.data:
                    metric_dict = {
                        'metric': metric.metric if hasattr(metric, 'metric') else None,
                        'description': metric.description if hasattr(metric, 'description') else None,
                        'data_type': metric.data_type if hasattr(metric, 'data_type') else None,
                        'category': category
                    }
                    if metric_dict['metric']:  # Only add if metric code exists
                        metrics_list.append(metric_dict)
                
                all_metrics[category] = metrics_list
                count = len(metrics_list)
                total_metrics += count
                logger.info(f"    ✅ Found {count} metrics in {category}")
            else:
                all_metrics[category] = []
                logger.warning(f"    ⚠️ No metrics found for {category}")
            
            time.sleep(0.5)  # Rate limiting
            
        except Exception as e:
            logger.error(f"    ❌ Error fetching {category}: {str(e)}")
            all_metrics[category] = []
    
    logger.info(f"📊 Total metrics discovered: {total_metrics}")
    return all_metrics

def get_metric_value_for_bank(
    api_client,
    bank_ticker: str,
    metrics_batch: List[str],
    fiscal_year: int = TARGET_FISCAL_YEAR,
    fiscal_quarter: int = TARGET_FISCAL_QUARTER
) -> Dict[str, Any]:
    """Get metric values for a specific bank for Q1 2025."""
    
    fund_api = fact_set_fundamentals_api.FactSetFundamentalsApi(api_client)
    
    # Use specific date range for Q1 2025
    start_date = f"{fiscal_year}-01-01"
    end_date = f"{fiscal_year}-03-31"
    
    try:
        # Create request
        ids_instance = IdsBatchMax30000([bank_ticker])
        metrics_instance = Metrics(metrics_batch)
        periodicity_instance = Periodicity("QTR")
        update_type_instance = UpdateType("RP")
        fiscal_period_instance = FiscalPeriod(
            start=start_date,
            end=end_date
        )
        batch_instance = Batch("N")
        
        request_data = FundamentalRequestBody(
            ids=ids_instance,
            metrics=metrics_instance,
            periodicity=periodicity_instance,
            fiscal_period=fiscal_period_instance,
            currency="CAD",  # Standardize to CAD for comparison
            update_type=update_type_instance,
            batch=batch_instance
        )
        
        request = FundamentalsRequest(data=request_data)
        
        # Make API call
        response_wrapper = fund_api.get_fds_fundamentals_for_list(request)
        
        # Unwrap response
        if hasattr(response_wrapper, 'get_response_200'):
            response = response_wrapper.get_response_200()
        else:
            response = response_wrapper
        
        # Process response
        metric_values = {}
        if response and hasattr(response, 'data') and response.data:
            for item in response.data:
                if hasattr(item, 'metric') and hasattr(item, 'value'):
                    # Check if value is not None and fiscal period matches Q1 2025
                    if item.value is not None:
                        fiscal_year_match = getattr(item, 'fiscal_year', None) == fiscal_year
                        fiscal_period_match = getattr(item, 'fiscal_period', None) == fiscal_quarter
                        
                        # Store value if it's from Q1 2025 or if no period info (latest available)
                        if fiscal_year_match and fiscal_period_match:
                            metric_values[item.metric] = {
                                'value': item.value,
                                'fiscal_year': fiscal_year,
                                'fiscal_period': fiscal_quarter,
                                'date': getattr(item, 'fiscal_end_date', None)
                            }
                        elif item.metric not in metric_values:
                            # Use latest available if Q1 2025 not found
                            metric_values[item.metric] = {
                                'value': item.value,
                                'fiscal_year': getattr(item, 'fiscal_year', None),
                                'fiscal_period': getattr(item, 'fiscal_period', None),
                                'date': getattr(item, 'fiscal_end_date', None)
                            }
        
        return metric_values
        
    except Exception as e:
        logger.debug(f"Error fetching metrics for {bank_ticker}: {str(e)}")
        return {}

def build_coverage_matrix(
    api_client,
    all_metrics: Dict[str, List[Dict[str, Any]]],
    banks: Dict[str, Dict[str, str]]
) -> pd.DataFrame:
    """Build comprehensive coverage matrix for all banks."""
    
    logger.info("🔨 Building coverage matrix...")
    
    # Prepare data structure
    rows = []
    bank_tickers = list(banks.keys())
    
    # Process each category and metric
    for category, metrics in all_metrics.items():
        if not metrics:
            continue
        
        logger.info(f"\n📊 Processing {category} ({len(metrics)} metrics)")
        
        # Group metrics by data type for efficient API calls
        metrics_by_type = {}
        metric_info = {}
        
        for metric in metrics:
            metric_code = metric['metric']
            data_type = metric.get('data_type', 'unknown')
            
            if data_type not in metrics_by_type:
                metrics_by_type[data_type] = []
            metrics_by_type[data_type].append(metric_code)
            metric_info[metric_code] = metric
        
        # Process each bank
        bank_data = {}
        for bank_ticker in bank_tickers:
            bank_name = banks[bank_ticker]['name']
            logger.info(f"  🏦 Checking {bank_ticker} ({bank_name})...")
            
            bank_metrics = {}
            
            # Test metrics in batches by data type
            for data_type, metric_codes in metrics_by_type.items():
                # Process in batches of 20
                for i in range(0, len(metric_codes), 20):
                    batch = metric_codes[i:i+20]
                    
                    # Get values for this batch
                    values = get_metric_value_for_bank(api_client, bank_ticker, batch)
                    bank_metrics.update(values)
                    
                    time.sleep(0.3)  # Rate limiting
            
            bank_data[bank_ticker] = bank_metrics
            logger.info(f"    ✅ Found data for {len(bank_metrics)} metrics")
        
        # Create rows for each metric
        for metric_code, info in metric_info.items():
            row = {
                'Category': category,
                'Metric Code': metric_code,
                'Description': info.get('description', ''),
                'Data Type': info.get('data_type', ''),
                'Period': f"FY{TARGET_FISCAL_YEAR} Q{TARGET_FISCAL_QUARTER}"
            }
            
            # Add bank values
            banks_with_data = 0
            for bank_ticker in bank_tickers:
                if metric_code in bank_data.get(bank_ticker, {}):
                    value_info = bank_data[bank_ticker][metric_code]
                    row[bank_ticker] = value_info['value']
                    banks_with_data += 1
                else:
                    row[bank_ticker] = None
            
            # Add analysis columns
            row['Banks with Data'] = banks_with_data
            row['Any Bank Has Data'] = 'Yes' if banks_with_data > 0 else 'No'
            row['All Banks Have Data'] = 'Yes' if banks_with_data == len(bank_tickers) else 'No'
            row['Coverage %'] = round((banks_with_data / len(bank_tickers)) * 100, 1)
            
            rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Sort by coverage percentage (descending) and category
    df = df.sort_values(['Coverage %', 'Category', 'Metric Code'], ascending=[False, True, True])
    
    return df

def format_excel_output(df: pd.DataFrame, banks: Dict[str, Dict[str, str]], output_path: str):
    """Create formatted Excel file with coverage matrix."""
    
    logger.info(f"📝 Creating Excel output: {output_path}")
    
    # Create Excel writer with xlsxwriter engine for better formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Write main data sheet
        df.to_excel(writer, sheet_name='Coverage Matrix', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Coverage Matrix']
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        bank_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        analysis_header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if col_name in ['Category', 'Metric Code', 'Description', 'Data Type', 'Period']:
                cell.fill = header_fill
                cell.font = header_font
            elif col_name in banks.keys():
                cell.fill = bank_header_fill
                cell.font = Font(bold=True)
            else:  # Analysis columns
                cell.fill = analysis_header_fill
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes (freeze first row and first 5 columns)
        worksheet.freeze_panes = 'F2'
        
        # Add conditional formatting for coverage percentage
        from openpyxl.formatting.rule import ColorScaleRule
        coverage_col = df.columns.get_loc('Coverage %') + 1
        worksheet.conditional_formatting.add(
            f'{get_column_letter(coverage_col)}2:{get_column_letter(coverage_col)}{len(df) + 1}',
            ColorScaleRule(
                start_type='min', start_color='FF0000',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='00FF00'
            )
        )
        
        # Create summary sheet
        summary_df = create_summary_sheet(df, banks)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        for col_num, col_name in enumerate(summary_df.columns, 1):
            cell = summary_sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust summary columns
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    logger.info(f"✅ Excel file created: {output_path}")

def create_summary_sheet(df: pd.DataFrame, banks: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    """Create summary statistics sheet."""
    
    bank_cols = list(banks.keys())
    
    # Overall statistics
    overall_stats = {
        'Metric': ['Total Metrics', 'Metrics with Any Data', 'Metrics with All Banks', 
                   'Average Coverage %', 'Median Coverage %'],
        'Value': [
            len(df),
            len(df[df['Any Bank Has Data'] == 'Yes']),
            len(df[df['All Banks Have Data'] == 'Yes']),
            round(df['Coverage %'].mean(), 1),
            round(df['Coverage %'].median(), 1)
        ]
    }
    
    # Category breakdown
    category_stats = []
    for category in df['Category'].unique():
        cat_df = df[df['Category'] == category]
        category_stats.append({
            'Category': category,
            'Total Metrics': len(cat_df),
            'With Data': len(cat_df[cat_df['Any Bank Has Data'] == 'Yes']),
            'Full Coverage': len(cat_df[cat_df['All Banks Have Data'] == 'Yes']),
            'Avg Coverage %': round(cat_df['Coverage %'].mean(), 1)
        })
    
    # Bank coverage statistics
    bank_stats = []
    for bank_ticker in bank_cols:
        bank_name = banks[bank_ticker]['name']
        bank_type = banks[bank_ticker]['type']
        
        # Count non-null values for this bank
        bank_coverage = df[bank_ticker].notna().sum()
        bank_percentage = round((bank_coverage / len(df)) * 100, 1)
        
        bank_stats.append({
            'Bank Ticker': bank_ticker,
            'Bank Name': bank_name,
            'Bank Type': bank_type,
            'Metrics Available': bank_coverage,
            'Coverage %': bank_percentage
        })
    
    # Create summary DataFrames
    overall_df = pd.DataFrame(overall_stats)
    category_df = pd.DataFrame(category_stats)
    bank_df = pd.DataFrame(bank_stats)
    
    # Sort bank statistics by coverage
    bank_df = bank_df.sort_values('Coverage %', ascending=False)
    
    # Combine into single sheet with spacing
    summary_parts = []
    summary_parts.append(pd.DataFrame({'Section': ['OVERALL STATISTICS'], 'Value': ['']}))
    summary_parts.append(overall_df)
    summary_parts.append(pd.DataFrame({'': ['', '']}))  # Spacer
    summary_parts.append(pd.DataFrame({'Section': ['CATEGORY BREAKDOWN'], 'Value': ['']}))
    summary_parts.append(category_df)
    summary_parts.append(pd.DataFrame({'': ['', '']}))  # Spacer
    summary_parts.append(pd.DataFrame({'Section': ['BANK COVERAGE'], 'Value': ['']}))
    summary_parts.append(bank_df)
    
    # Concatenate all parts
    summary_df = pd.concat(summary_parts, ignore_index=True)
    
    return summary_df

def main():
    """Main function to generate coverage matrix."""
    logger.info("="*80)
    logger.info("🏦 FACTSET FUNDAMENTALS COVERAGE MATRIX GENERATOR")
    logger.info("="*80)
    logger.info(f"Target Period: Q{TARGET_FISCAL_QUARTER} {TARGET_FISCAL_YEAR}")
    logger.info("="*80)
    
    # Validate environment
    validate_env_vars()
    
    # Load configuration
    config = load_config_yaml()
    
    # Get Canadian and US banks only
    banks = get_canadian_and_us_banks(config)
    logger.info(f"📊 Analyzing {len(banks)} banks:")
    for ticker, info in list(banks.items())[:5]:
        logger.info(f"  • {ticker}: {info['name']} ({info['type']})")
    if len(banks) > 5:
        logger.info(f"  ... and {len(banks) - 5} more")
    
    # Setup SSL certificate
    ssl_cert_path = setup_ssl_certificate()
    if not ssl_cert_path:
        logger.warning("⚠️ No SSL certificate available - continuing without SSL verification")
    
    # Configure API
    configuration = setup_api_configuration(ssl_cert_path)
    
    try:
        with fds.sdk.FactSetFundamentals.ApiClient(configuration) as api_client:
            
            # Phase 1: Get all available metrics
            logger.info("\n📊 PHASE 1: Discovering all available metrics")
            logger.info("-"*60)
            all_metrics = get_all_available_metrics(api_client)
            
            # Phase 2: Build coverage matrix
            logger.info("\n📊 PHASE 2: Building coverage matrix for all banks")
            logger.info("-"*60)
            
            df = build_coverage_matrix(api_client, all_metrics, banks)
            
            # Phase 3: Generate outputs
            logger.info("\n📊 PHASE 3: Generating outputs")
            logger.info("-"*60)
            
            # Save to CSV (backup)
            csv_filename = f"coverage_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            df.to_csv(csv_filename, index=False)
            logger.info(f"✅ CSV backup saved: {csv_filename}")
            
            # Create formatted Excel
            excel_filename = f"FactSet_Fundamentals_Coverage_Matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            format_excel_output(df, banks, excel_filename)
            
            # Print summary statistics
            logger.info("\n" + "="*80)
            logger.info("📊 COVERAGE SUMMARY")
            logger.info("="*80)
            
            total_metrics = len(df)
            metrics_with_data = len(df[df['Any Bank Has Data'] == 'Yes'])
            full_coverage = len(df[df['All Banks Have Data'] == 'Yes'])
            avg_coverage = df['Coverage %'].mean()
            
            logger.info(f"Total Metrics Analyzed: {total_metrics}")
            logger.info(f"Metrics with Any Bank Data: {metrics_with_data} ({metrics_with_data/total_metrics*100:.1f}%)")
            logger.info(f"Metrics with Full Coverage: {full_coverage} ({full_coverage/total_metrics*100:.1f}%)")
            logger.info(f"Average Coverage: {avg_coverage:.1f}%")
            
            # Top covered metrics
            top_metrics = df.nlargest(10, 'Coverage %')[['Metric Code', 'Description', 'Coverage %']]
            logger.info("\n📈 Top 10 Best Covered Metrics:")
            for _, row in top_metrics.iterrows():
                logger.info(f"  • {row['Metric Code']}: {row['Coverage %']}% - {row['Description'][:50]}...")
            
            logger.info("\n✅ Analysis complete!")
            logger.info(f"📊 View the Excel file for detailed analysis: {excel_filename}")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {str(e)}")
        raise
    
    finally:
        # Cleanup
        if ssl_cert_path and ssl_cert_path.startswith('/tmp/'):
            try:
                os.unlink(ssl_cert_path)
            except:
                pass

if __name__ == "__main__":
    main()